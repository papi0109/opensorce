#!/usr/bin/python3
# _*_ coding: utf-8 _*_

import re
import os
import openpyxl as op
from pyparsing import line
import collections as cl
import configparser as cfp
import logging as log
from tkinter import filedialog
import argparse

parser = argparse.ArgumentParser(
    description='ログ内の文字列を抽出するスクリプトです。')
parser.add_argument('-s','--select',action='store_true',help='ファイル選択ダイアログ表示')
args = parser.parse_args()

outputFile = "./output.xlsx"

log.basicConfig(
    filename="./extract.log",
    level=log.DEBUG,
    format="%(asctime)s [%(levelname)s] (PID:%(process)d)) %(message)s",
    datefmt="%Y/%m/%d %H:%M:%S"
)

configname = "config.ini"
current = os.path.dirname(__file__) + "/"
config  = cfp.ConfigParser()
config.read(current+configname,encoding="utf-8")
regexString = config["regex"]["var"]

try:
    if args.select:
        filetype = [('ログファイル','*.log'),('テキストファイル','*.txt')] 
        readingFile = filedialog.askopenfilename(filetypes = filetype) 
        if not readingFile:
            exit(0)
        else:
            pass
        print(readingFile)
        print(type(readingFile))
    else:
        readingFile = config["logfile"]["log"]
        foundFile:bool = os.path.isfile(readingFile)
        if foundFile:
            msg = "ファイルの存在を確認できました。処理を続けます。"
            log.info(msg)
        else:
            msg = f"指定したログファイルが存在しません。{readingFile} を設置し、スタートし直して下さい。"
            log.error(msg)
            exit(1)
            
except FileNotFoundError:
    msg = "指定したファイルが存在しません。"
    log.exception(msg)
    exit(1)

countAll = 0
with open(readingFile,mode="r",encoding="utf-8") as f:
    lines = f.readlines()
    test_list = []
    for Log_file in lines:
        Extract = re.findall(regexString, Log_file)
        test_list += Extract
        lineCount = len(test_list)

        if lineCount == 0:
            break
        else:
            try:
                msg = f"{lineCount} 回目 ( {test_list[-1]} ) を抽出"
                log.info(msg)
                countAll += 1
                
            except IndexError:
                break
            
            except Exception:
                msg = "抽出処理に失敗しました。"
                log.exception(msg)
                exit(1)

duplicate:dict = cl.Counter(test_list)
duplicate      = [k for k, v in duplicate.items() if v > 1]
duplicateCount = len(duplicate)

noDupList  = list((dict.fromkeys(test_list)))
noDupCount = len(noDupList)

if not countAll:
    msg = "抽出することが出来ませんでした。抽出箇所と正規表現を変えてみてください。"
    log.error(msg)
    exit(1)
else:
    pass

msg = f"{len(noDupList)} 名の抽出が完了しました。"
log.info(msg)

if duplicateCount > 0:
    log.info("重複していたユーザーは以下です。")
    for i in range(duplicateCount):
        log.info(f"({i+1}) {duplicate[i]}")
    
    msg = f"{duplicateCount} 名が重複していたことを確認しました。"
    log.info(msg)
else:
    pass

def createSheet(exelfile):
    exFile = op.Workbook()
    exFile.save(exelfile)

is_file = os.path.isfile(outputFile)
if is_file:
    msg = f"{outputFile} が存在する為、作成し直してから出力します。"
    os.remove(outputFile)
    createSheet(outputFile)
else:
    msg = f"\n[INFO] {outputFile} が存在しない為、作成します。"
    createSheet(outputFile)
    
log.info(msg)

wb = op.load_workbook(outputFile)
ws = wb.active
exeCount = 0
while exeCount <= noDupCount:
    try:
        ws.cell(exeCount + 1, 1, value=noDupList[exeCount])
        msg = f"セル A{exeCount} にデータ[ {noDupList[exeCount]} ]の入力完了"
        print(msg)
        log.info(msg)
    
        wb.save(outputFile)
        exeCount = exeCount+1
        
    except FileNotFoundError:
        createSheet(outputFile)
        wb.save(outputFile)
    
    except IndexError:
        break
    
    except Exception:
        msg = "データ入力処理でエラーが発生しました。"
        log.exception(msg)
        break

log.info(f"処理が完了しました。")
exit(0)
